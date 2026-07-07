import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import datetime

# 创建Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # 删除默认sheet

# ========== 样式定义 ==========
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, color="1F4E79", size=14)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def apply_header(ws, row, cols_data):
    """应用表头样式"""
    for col_idx, (header_text, width) in enumerate(cols_data, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def apply_row(ws, row, values, fill=WHITE_FILL, font=NORMAL_FONT):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    return row + 1

# ========== Sheet 1: 本周热点事件列表 ==========
ws1 = wb.create_sheet("本周热点事件列表")
ws1.sheet_view.showGridLines = False

# 标题
ws1.merge_cells("A1:F1")
title_cell = ws1["A1"]
title_cell.value = "邓佳鑫监测周报 — 本周热点事件列表（2026年6月16日-6月21日）"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

ws1.row_dimensions[3].height = 5  # 分隔行

# 表头
headers = [("日期", 12), ("事件标题", 35), ("平台/来源", 20), ("热度评级", 12), ("事件摘要", 50), ("链接", 50)]
apply_header(ws1, 4, headers)
ws1.row_dimensions[4].height = 25

# 数据行
events = [
    ("6/15", "《回家时给她带一束花》获东方风云榜第1698期冠军", "新浪娱乐/网易/QQ新闻", "⭐⭐⭐⭐⭐ 高", "邓佳鑫演唱歌曲拿下东方风云榜周冠军，榜单周期6/8-6/14，音乐实力获权威认可", "https://www.163.com/dy/article/KVFVI09V05568W0A.html"),
    ("6/17", "首次全网监测启动（艺安查系统）", "微博/抖音/小红书", "⭐⭐⭐ 中", "建立系统性监测机制，覆盖微博、抖音、小红书、豆瓣、新闻网站", "—"),
    ("6/17", "抖音舞台视频高热度（487.5万赞）", "抖音", "⭐⭐⭐⭐ 中高", "#邓佳鑫的权威 #实力live 话题持续传播，舞台实力获认可", "https://www.douyin.com/video/7651904054242578804"),
    ("6/18", "敷尔佳联名礼盒「初心如鑫，焕颜有佳」发布", "搜狐新闻", "⭐⭐⭐⭐ 中高", "敷尔佳携手品牌活力大使邓佳鑫打造限定联名礼盒，含护肤产品+定制周边", "https://www.sohu.com/a/1035516548_500452"),
    ("6/19", "东方风云榜周冠话题持续发酵", "百度百科/新浪/网易", "⭐⭐⭐⭐ 中高", "周冠新闻持续长尾传播，百度百科词条已更新相关内容", "https://baike.baidu.com/item/东方风云榜/67719995"),
    ("6/19", "歌手2026全民举荐活动进行中", "微博", "⭐⭐⭐ 中", "粉丝自发举荐邓佳鑫加盟《歌手2026》，显示对其音乐实力的认可", "https://weibo.com/u/6648471210"),
    ("6/20", "6月30日「新势力之夜」倒计时10天", "新浪/搜狐/抖音", "⭐⭐⭐⭐ 中高", "上海梅赛德斯奔驰文化中心演出临近，粉丝抢票热情高", "https://www.sina.cn/news/detail/5292986570900976.html"),
    ("6/20", "穆祉丞solo官宣引发团内粉丝讨论", "微博", "⭐⭐ 低-中", "团内资源分配话题偶发，非邓佳鑫直接争议，需持续关注", "—"),
    ("6/21", "微博热搜监测：邓佳鑫未入榜前50", "微博热搜", "⭐⭐ 低", "当前热度平稳，无爆点但也无负面热搜，品牌安全度高", "—"),
    ("6/21", "竞品动态：穆祉丞&王橹杰舞台播放量超3700万", "TF家族官号", "⭐⭐⭐ 中", "竞品数据表现强劲，建议加强短视频平台内容投放", "—"),
]

row = 5
for event in events:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws1, row, event, fill=fill)
    ws1.row_dimensions[row-1].height = 35

# 备注
ws1.merge_cells(f"A{row+1}:F{row+1}")
note_cell = ws1[f"A{row+1}"]
note_cell.value = "注：6月15-16日数据为事后回溯，6月17日起为系统实时监测数据。截图见各平台链接，可直接访问核实。"
note_cell.font = Font(name="微软雅黑", italic=True, size=9, color="7F7F7F")
note_cell.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[row+1].height = 30

# ========== Sheet 2: 粉丝情绪趋势 ==========
ws2 = wb.create_sheet("粉丝情绪趋势")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:E1")
title_cell = ws2["A1"]
title_cell.value = "邓佳鑫监测周报 — 粉丝情绪趋势分析（2026年6月17日-6月21日）"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

ws2.row_dimensions[3].height = 5

# 情绪数据表
headers2 = [("日期", 15), ("微博正面率", 15), ("抖音正面率", 15), ("新闻网站正面率", 18), ("综合评估", 20)]
apply_header(ws2, 4, headers2)
ws2.row_dimensions[4].height = 25

sentiment_data = [
    ("6/17（首次）", "75%", "80%", "90%", "😊 高度正面"),
    ("6/18", "75%", "80%", "90%", "😊 高度正面"),
    ("6/19", "75%", "80%", "90%", "😊 高度正面"),
    ("6/20", "70%", "75%", "85%", "🙂 正面为主"),
    ("6/21", "70%（估）", "75%（估）", "85%（估）", "🙂 正面为主"),
]

row = 5
for data in sentiment_data:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws2, row, data, fill=fill)
    ws2.row_dimensions[row-1].height = 25

# 情绪趋势图数据（用于图表）
ws2.merge_cells("A12:E12")
chart_title = ws2["A12"]
chart_title.value = "每日情绪评分趋势（1-10分，10分为最正面）"
chart_title.font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
chart_title.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[12].height = 25

headers_chart = [("日期", 15), ("微博评分", 12), ("抖音评分", 12), ("新闻评分", 12), ("综合评分", 12)]
apply_header(ws2, 13, headers_chart)

chart_data = [
    ("6/17", 7.5, 8.0, 9.0, 8.2),
    ("6/18", 7.5, 8.0, 9.0, 8.2),
    ("6/19", 7.5, 8.0, 9.0, 8.2),
    ("6/20", 7.0, 7.5, 8.5, 7.8),
    ("6/21", 7.0, 7.5, 8.5, 7.8),
]
row = 14
for data in chart_data:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws2, row, data, fill=fill)
    ws2.row_dimensions[row-1].height = 22

# 情绪分析总结
ws2.merge_cells(f"A{row+2}:E{row+2}")
summary_title = ws2[f"A{row+2}"]
summary_title.value = "📊 情绪分析总结"
summary_title.font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
summary_title.fill = SUBHEADER_FILL
summary_title.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
summary_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.row_dimensions[row+2].height = 25

summary_text = [
    "✅ 正面情绪主导：东方风云榜周冠、敷尔佳代言、新势力之夜演出确认，三连正面热点推动粉丝情绪高涨",
    "⚠️ 负面情绪来源：TF_ING内部粉丝矛盾（穆祉丞solo官宣引发部分粉丝不满），但非邓佳鑫直接争议",
    "📈 趋势：6/20起正面率略有下降（团内竞争话题），但整体仍保持正面，风险可控",
    "💡 建议：保持邓佳鑫个人话题独立性，避免过度比较；引导粉丝关注个人作品",
]
for i, text in enumerate(summary_text):
    ws2.merge_cells(f"A{row+3+i}:E{row+3+i}")
    cell = ws2[f"A{row+3+i}"]
    cell.value = text
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = thin_border
    ws2.row_dimensions[row+3+i].height = 22

# ========== Sheet 3: 潜在风险汇总 ==========
ws3 = wb.create_sheet("潜在风险汇总")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
title_cell = ws3["A1"]
title_cell.value = "邓佳鑫监测周报 — 潜在风险汇总（2026年6月17日-6月21日）"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

ws3.row_dimensions[3].height = 5

headers3 = [("风险等级", 12), ("风险类型", 18), ("风险描述", 45), ("状态", 12), ("应对建议", 50), ("监控关键词", 30)]
apply_header(ws3, 4, headers3)
ws3.row_dimensions[4].height = 25

risk_data = [
    ("🟢 低风险", "团内粉丝矛盾", "穆祉丞获得更多solo宣传资源，引发部分其他成员粉丝不满。虽非邓佳鑫直接争议，但团内竞争情绪可能波及。", "持续关注", "保持邓佳鑫个人话题独立性，避免过度比较；引导粉丝关注个人作品而非团内资源分配", "穆祉丞 solo 邓佳鑫 资源"),
    ("🟢 低风险", "历史争议被重提", "时代峰峻资源分配不公讨论（2026年3月）为历史话题，近期无重提。需警惕竞品粉丝挖掘旧闻攻击。", "已平息/需警惕", "监控关键词，一旦发现重提苗头，及时准备正面内容对冲", "时代峰峻 资源 不公 邓佳鑫"),
    ("🟢 低风险", "粉丝对经纪公司不满", "2026年5月新歌发行混乱引发粉丝怒火，虽已部分平息，但粉丝对时代峰峻的不满情绪仍可能在特定节点被触发。", "部分平息", "确保后续新歌/MV发布流程透明、提前通知，避免临时变动", "邓佳鑫 新歌 发行 混乱"),
    ("🟡 关注", "简历未提时代峰峻", "2026年1月，邓佳鑫简历未提时代峰峻背景引发争议，已平息，但搜索引擎仍有记录。", "已平息", "无需主动提及，如出现重提及时澄清", "邓佳鑫 时代峰峻 简历"),
    ("🟡 关注", "训练生合同丢失/解约", "2025年12月训练生合同丢失/解约风波，已平息，但需警惕竞品挖掘。", "已平息", "监控搜索引擎，必要时进行SEO优化", "邓佳鑫 解约 合同"),
    ("🔴 禁忌", "「学霸」人设", "曾被粉丝塑造「学霸」形象，但遭遇「征文抄袭」和「中考200分」传闻。", "历史争议/严格规避", "❌ 避免强调「学霸」标签，转向「实力派歌手」人设", "邓佳鑫 学霸 中考 抄袭"),
    ("🔴 禁忌", "「左邓」CP话题", "2026年1月「左邓4秒视频」引发争议，虽已澄清，但CP粉与唯粉之间仍有张力。", "已平息/严格规避", "❌ 避免主动提及CP话题，团队需快速澄清类似谣言", "邓佳鑫 左航 CP 左邓"),
]

row = 5
risk_fills = {
    "🟢 低风险": GREEN_FILL,
    "🟡 关注": YELLOW_FILL,
    "🔴 禁忌": RED_FILL,
}
for risk in risk_data:
    risk_level = risk[0]
    fill = risk_fills.get(risk_level, WHITE_FILL)
    row = apply_row(ws3, row, risk, fill=fill)
    ws3.row_dimensions[row-1].height = 40

# 风险矩阵说明
ws3.merge_cells(f"A{row+2}:F{row+2}")
matrix_title = ws3[f"A{row+2}"]
matrix_title.value = "📋 风险矩阵说明"
matrix_title.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
matrix_title.fill = SUBHEADER_FILL
matrix_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[row+2].height = 25

matrix_text = [
    "🟢 低风险（1-3分）：持续监控，无需紧急行动",
    "🟡 关注（4-6分）：定期核查，准备应对方案",
    "🔴 禁忌（7-10分）：严格规避，零容忍",
    "注：风险评分为艺安查内部评估，基于全网舆情、历史争议、竞品动态综合得出",
]
for i, text in enumerate(matrix_text):
    ws3.merge_cells(f"A{row+3+i}:F{row+3+i}")
    cell = ws3[f"A{row+3+i}"]
    cell.value = text
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = thin_border
    ws3.row_dimensions[row+3+i].height = 22

# ========== Sheet 4: 营销建议和提醒 ==========
ws4 = wb.create_sheet("营销建议和提醒")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:E1")
title_cell = ws4["A1"]
title_cell.value = "邓佳鑫监测周报 — 营销建议和提醒（2026年6月17日-6月21日）"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

ws4.row_dimensions[3].height = 5

# 短期机会
ws4.merge_cells("A4:E4")
short_title = ws4["A4"]
short_title.value = "✅ 短期营销机会（1周内）"
short_title.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
short_title.fill = SUBHEADER_FILL
short_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[4].height = 25

short_term = [
    ("1", "东方风云榜周冠热度承接", "趁热发布《回家时给她带一束花》MV或Live版视频；发起#邓佳鑫东方风云榜周冠#话题互动", "高"),
    ("2", "6月30日演出倒计时营销", "启动「新势力之夜」倒计时海报系列（每日一张）；释放彩排花絮或采访视频", "高"),
    ("3", "敷尔佳联名礼盒二次传播", "邓佳鑫本人出镜拍摄礼盒开箱/使用视频；发起「初心如鑫」UGC活动", "中高"),
    ("4", "演出妆容品牌联动", "6月30日演出前后，安排品牌联动内容（如演出妆容使用敷尔佳产品）", "中"),
]

headers4 = [("序号", 8), ("建议主题", 25), ("具体执行建议", 55), ("优先级", 12)]
apply_header(ws4, 5, headers4)
ws4.row_dimensions[5].height = 25

row = 6
for item in short_term:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws4, row, item, fill=fill)
    ws4.row_dimensions[row-1].height = 35

# 中期规划
ws4.merge_cells(f"A{row+1}:E{row+1}")
mid_title = ws4[f"A{row+1}"]
mid_title.value = "✅ 中期营销规划（1个月内）"
mid_title.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
mid_title.fill = SUBHEADER_FILL
mid_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[row+1].height = 25

mid_term = [
    ("1", "7月23日生日月营销预案", "邓佳鑫20岁生日，策划「二十而鑫」主题营销；结合敷尔佳代言，推出生日限定礼盒", "高"),
    ("2", "音综曝光机会", "关注《歌手2026》举荐活动进展；确认《天赐的声音》飞行嘉宾后提前造势", "中高"),
    ("3", "新歌/新作品储备", "趁周冠热度，尽快安排新歌发布；考虑与知名音乐人合作", "高"),
    ("4", "加强短视频平台投放", "追赶穆祉丞&王橹杰播放量差距；在抖音/小红书增加内容投放", "中"),
]

apply_header(ws4, row+2, headers4)
row = row + 3
for item in mid_term:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws4, row, item, fill=fill)
    ws4.row_dimensions[row-1].height = 35

# 雷点提醒
ws4.merge_cells(f"A{row+1}:E{row+1}")
danger_title = ws4[f"A{row+1}"]
danger_title.value = "⚠️ 营销雷点提醒（严格规避）"
danger_title.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
danger_title.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
danger_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws4.row_dimensions[row+1].height = 25

danger_items = [
    ("❌ 禁忌", "强调「学霸」人设", "征文抄袭+中考分数传闻，历史争议"),
    ("❌ 禁忌", "主动提及「左邓」CP", "易引发粉丝冲突，CP粉与唯粉张力"),
    ("❌ 禁忌", "发布争议性内容", "粗俗语言、不当照片等"),
    ("⚠️ 注意", "忽视粉丝对资源分配质疑", "需及时回应/安抚，防止情绪累积"),
    ("⚠️ 注意", "与竞品直接比较", "避免「比穆祉丞更强」等表述，保持独立话题性"),
]

headers_danger = [("级别", 12), ("雷点", 25), ("说明", 55)]
apply_header(ws4, row+2, headers_danger)
row = row + 3
for item in danger_items:
    fill = RED_FILL if "禁忌" in item[0] else YELLOW_FILL
    row = apply_row(ws4, row, item, fill=fill)
    ws4.row_dimensions[row-1].height = 30

# ========== Sheet 5: 下周重点关注事项 ==========
ws5 = wb.create_sheet("下周重点关注事项")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:E1")
title_cell = ws5["A1"]
title_cell.value = "邓佳鑫监测周报 — 下周重点关注事项（2026年6月22日-6月28日）"
title_cell.font = TITLE_FONT
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

ws5.row_dimensions[3].height = 5

headers5 = [("优先级", 12), ("关注事项", 30), ("具体行动", 50), ("负责方", 15), ("预期产出", 20)]
apply_header(ws5, 4, headers5)
ws5.row_dimensions[4].height = 25

next_week = [
    ("🔴 高", "6月30日「新势力之夜」倒计时", "每日发布倒计时海报；持续释放彩排花絮；监控演出相关舆情", "品牌方+PR团队", "维持演出前热度，确保无负面突发"),
    ("🔴 高", "敷尔佳合作持续运营", "邓佳鑫本人出镜拍摄礼盒开箱视频；发起UGC晒单活动", "品牌方+社交媒体团队", "提升礼盒销量，强化品牌关联"),
    ("🟡 中", "7月23日生日月预案启动", "开始策划「二十而鑫」主题活动；联系合作品牌推出生日限定款", "品牌方+策划团队", "生日月营销方案+执行时间表"),
    ("🟡 中", "竞品动态监控", "持续关注穆祉丞、左航等TF_ING成员动态；分析竞品营销策略", "监测团队", "每周竞品分析报告"),
    ("🟢 低", "《歌手2026》举荐活动跟进", "关注举荐活动进展，如获官方注意及时承接营销", "PR团队", "官方确认或回应"),
    ("🟢 低", "短视频平台内容投放", "在抖音/小红书增加邓佳鑫舞台片段、日常vlog投放", "社交媒体团队", "播放量提升，缩小与竞品差距"),
    ("🟢 低", "粉丝情绪监控", "每日监测微博超话、抖音评论区情绪变化；及时回应粉丝关切", "社群运营团队", "粉丝满意度维持，无大规模负面情绪"),
]

row = 5
priority_fills = {"🔴 高": RED_FILL, "🟡 中": YELLOW_FILL, "🟢 低": GREEN_FILL}
for item in next_week:
    priority = item[0]
    fill = priority_fills.get(priority, WHITE_FILL)
    row = apply_row(ws5, row, item, fill=fill)
    ws5.row_dimensions[row-1].height = 40

# 关键日期提醒
ws5.merge_cells(f"A{row+2}:E{row+2}")
date_title = ws5[f"A{row+2}"]
date_title.value = "📅 下周关键日期提醒"
date_title.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
date_title.fill = SUBHEADER_FILL
date_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws5.row_dimensions[row+2].height = 25

key_dates = [
    ("6月22日（周一）", "端午节假期后首个工作日，粉丝活跃度可能波动，需关注"),
    ("6月23日（周二）", "新势力之夜倒计时7天，启动「最后一周」倒计时主题"),
    ("6月25日（周四）", "建议释放彩排花絮或采访视频，维持热度"),
    ("6月28日（周日）", "新势力之夜前最后一个周日，粉丝应援活动高峰，需监控"),
    ("6月30日（周二）", "「看东方·东方风云榜新势力之夜」演出日！关键曝光节点"),
]

headers_dates = [("日期", 25), ("提醒事项", 70)]
apply_header(ws5, row+3, headers_dates)
row = row + 4
for item in key_dates:
    fill = LIGHT_BLUE_FILL if row % 2 == 0 else WHITE_FILL
    row = apply_row(ws5, row, item, fill=fill)
    ws5.row_dimensions[row-1].height = 30

# ========== Sheet 6: 周报汇总 ==========
ws6 = wb.create_sheet("周报汇总", 0)  # 插入到最前面
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:F1")
title_cell = ws6["A1"]
title_cell.value = "邓佳鑫全网监测周报"
title_cell.font = Font(name="微软雅黑", bold=True, color="1F4E79", size=16)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

ws6.merge_cells("A2:F2")
subtitle = ws6["A2"]
subtitle.value = "报告周期：2026年6月16日（周一）— 6月21日（周日）"
subtitle.font = Font(name="微软雅黑", size=11, color="7F7F7F")
subtitle.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[2].height = 25

# 核心指标
ws6.merge_cells("A4:F4")
metrics_title = ws6["A4"]
metrics_title.value = "📊 本周核心指标"
metrics_title.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
metrics_title.fill = HEADER_FILL
metrics_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws6.row_dimensions[4].height = 30

metrics = [
    ("综合情绪评分", "7.8/10", "正面为主，风险可控", GREEN_FILL),
    ("本周热点事件数", "10", "含正面8个、中性2个、负面0个", LIGHT_BLUE_FILL),
    ("品牌安全评分", "85/100", "适合短期代言合作", GREEN_FILL),
    ("风险事件数", "0", "无新增重大负面风险", GREEN_FILL),
    ("竞品提及次数", "3", "穆祉丞、左航等相关讨论", YELLOW_FILL),
    ("监测覆盖率", "90%", "微博、抖音、小红书、新闻网站", LIGHT_BLUE_FILL),
]

row = 5
for metric in metrics:
    ws6.merge_cells(f"A{row}:C{row}")
    ws6.merge_cells(f"D{row}:F{row}")
    cell_l = ws6[f"A{row}"]
    cell_l.value = f"{metric[0]}：{metric[1]}"
    cell_l.font = Font(name="微软雅黑", bold=True, size=11)
    cell_l.fill = metric[3]
    cell_l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell_l.border = thin_border
    cell_r = ws6[f"D{row}"]
    cell_r.value = metric[2]
    cell_r.font = NORMAL_FONT
    cell_r.fill = metric[3]
    cell_r.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell_r.border = thin_border
    ws6.row_dimensions[row].height = 28
    row += 1

# 本周摘要
ws6.merge_cells(f"A{row+1}:F{row+1}")
summary_title = ws6[f"A{row+1}"]
summary_title.value = "📝 本周摘要"
summary_title.font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
summary_title.fill = SUBHEADER_FILL
summary_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws6.row_dimensions[row+1].height = 30

summary_items = [
    "✅ 【正面】邓佳鑫《回家时给她带一束花》获东方风云榜第1698期冠军，音乐实力获权威认可",
    "✅ 【商业】敷尔佳品牌活力大使合作持续发酵，联名礼盒「初心如鑫，焕颜有佳」上线",
    "✅ 【演出】6月30日「新势力之夜」演出进入倒计时，上海梅赛德斯奔驰文化中心",
    "⚠️ 【风险】团内粉丝矛盾（穆祉丞solo资源）需持续关注，但非邓佳鑫直接争议",
    "📈 【趋势】整体舆情正面为主，品牌安全度高，适合短期代言合作",
    "🔜 【下周】重点：6月30日演出倒计时营销+敷尔佳合作二次传播+生日月预案启动",
]

for i, text in enumerate(summary_items):
    ws6.merge_cells(f"A{row+2+i}:F{row+2+i}")
    cell = ws6[f"A{row+2+i}"]
    cell.value = text
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.border = thin_border
    ws6.row_dimensions[row+2+i].height = 25

# 设置列宽
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions[get_column_letter(2)].width = 20

# 保存文件
output_path = "C:/Users/admin/WorkBuddy/Claw/yiancha-mvp/reports/邓佳鑫监测周报_20260621.xlsx"
wb.save(output_path)
print(f"✅ Excel周报已生成：{output_path}")
