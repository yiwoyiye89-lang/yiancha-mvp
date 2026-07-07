#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
邓佳鑫监测周报Excel生成脚本 - 2026年6月30日-7月5日（第27周）
生成时间：2026年7月5日
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter

# 创建Excel工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认sheet

# 定义样式
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F3864')
subtitle_font = Font(name='Arial', bold=True, size=12, color='2E75B6')
normal_font = Font(name='Arial', size=10)
warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
neutral_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
info_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

# ===== Sheet 1: 封面 =====
sheet_cover = wb.create_sheet('封面')
sheet_cover.merge_cells('A1:H1')
sheet_cover['A1'] = '邓佳鑫全网监测周报'
set_cell_style(sheet_cover['A1'], Font(name='Arial', bold=True, size=20, color='1F3864'), None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[1].height = 40

sheet_cover.merge_cells('A3:H3')
sheet_cover['A3'] = '2026年6月30日 - 7月5日（第27周）'
set_cell_style(sheet_cover['A3'], subtitle_font, None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[3].height = 30

sheet_cover.merge_cells('A5:H5')
sheet_cover['A5'] = '报告生成时间：2026年7月5日 13:55'
set_cell_style(sheet_cover['A5'], normal_font, None, Alignment(horizontal='center', vertical='center'))

sheet_cover.merge_cells('A7:H7')
sheet_cover['A7'] = '生成工具：WorkBuddy 自动化监测系统'
set_cell_style(sheet_cover['A7'], normal_font, None, Alignment(horizontal='center', vertical='center'))

# 添加品牌安全评分
sheet_cover.merge_cells('A10:C10')
sheet_cover['A10'] = '本周品牌安全评分'
set_cell_style(sheet_cover['A10'], subtitle_font, None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[10].height = 25

sheet_cover.merge_cells('A11:C11')
sheet_cover['A11'] = '76-82/100'
set_cell_style(sheet_cover['A11'], Font(name='Arial', bold=True, size=24, color='FF8C00'), None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[11].height = 40

sheet_cover.merge_cells('A13:H13')
sheet_cover['A13'] = '⚠️ 本周评级：中低风险（敷尔佳合作拉升，CP话题持续需关注）'
set_cell_style(sheet_cover['A13'], Font(name='Arial', bold=True, size=12, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[13].height = 30

# 本周关键数据
sheet_cover.merge_cells('A16:B16')
sheet_cover['A16'] = '本周关键数据'
set_cell_style(sheet_cover['A16'], subtitle_font, info_fill, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[16].height = 25

key_data = [
    ('微博粉丝', '554.6万'),
    ('V指数（文娱）', '84.34分'),
    ('敷尔佳开售24h战报', '240w+'),
    ('本周重大事件', '5个'),
    ('品牌合作', '敷尔佳活力大使'),
    ('下周年终节点', '生日（7/23）倒计时18天')
]

for i, (key, value) in enumerate(key_data):
    row = 17 + i
    sheet_cover[f'A{row}'] = key
    sheet_cover[f'B{row}'] = value
    set_cell_style(sheet_cover[f'A{row}'], normal_font, None, Alignment(horizontal='right', vertical='center'))
    set_cell_style(sheet_cover[f'B{row}'], Font(name='Arial', bold=True, size=11, color='1F3864'), None, Alignment(horizontal='left', vertical='center'))
    sheet_cover.row_dimensions[row].height = 20

# 调整封面sheet列宽
sheet_cover.column_dimensions['A'].width = 20
sheet_cover.column_dimensions['B'].width = 20
sheet_cover.column_dimensions['C'].width = 20

# ===== Sheet 2: 本周热点事件 =====
sheet_events = wb.create_sheet('本周热点事件')
sheet_events['A1'] = '本周热点事件列表（2026年6月30日-7月5日）'
sheet_events.merge_cells('A1:H1')
set_cell_style(sheet_events['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_events.row_dimensions[1].height = 30

# 表头
headers = ['序号', '事件名称', '热度等级', '发生时间', '情绪倾向', '风险等级', '相关链接', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_events.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_events.row_dimensions[2].height = 30

# 热点事件数据（本周实际发生）
events_data = [
    [1, '东方风云榜新势力之夜25分钟微个唱', '⭐⭐⭐⭐⭐', '2026-06-30', '正面85%', '低风险', 'https://baike.baidu.com/item/东方风云榜新势力之夜/67719995', '邓佳鑫最重要的个人舞台曝光，表现获好评'],
    [2, 'TF_ING全员黑西装红毯造型', '⭐⭐⭐⭐', '2026-06-29', '正面90%', '低风险', 'https://k.sina.cn/article_7879776328_1d5abd84806801sp9g.html', '团体形象升级，邓佳鑫获赞"周正大帅哥"'],
    [3, '敷尔佳品牌活力大使宣发启动', '⭐⭐⭐⭐⭐', '2026-07-03', '正面88%', '低风险', 'https://m.163.com/news/article/KVB4KHD705118CEK.html', '7/3-7/10全程宣发，TVC花絮+直播+限定礼盒'],
    [4, '"左邓吵架"CP话题波动', '⭐⭐⭐⭐', '2026-07-01', '两极分化', '高风险', 'https://www.douyin.com/video/7657545086414462450', '抖音4.9万赞，需防范影响敷尔佳合作'],
    [5, '穆祉丞"躺照"争议事件', '⭐⭐⭐⭐', '2026-06-29', '负面', '中风险', 'https://ent.sina.cn/2026-06-30/detail-inifcvuh6720269.d.html', 'TF_ING内部矛盾公开化，可能波及邓佳鑫'],
    [6, '穆祉丞常州站演唱会倒计时', '⭐⭐⭐', '2026-07-11', '中性', '中风险', 'https://ent.sina.cn/2026-06-27/detail-inievzuq4607123.d.html', '与敷尔佳直播期重叠，声量可能被分流'],
    [7, 'TF家族夏日运动会官宣（澳门）', '⭐⭐⭐⭐', '2026-07-16', '期待70%', '中风险', 'https://ent.sina.cn/2026-06-28/detail-iniexymu0628039.d.html', '运动会为CP话题高发场景，需提前准备'],
]

for row_idx, row_data in enumerate(events_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_events.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据风险等级设置颜色
        if col_idx == 6:  # 风险等级列
            if '高' in str(value):
                set_cell_style(cell, normal_font, warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, normal_font, neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '低' in str(value):
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True))
        
        cell.border = thin_border
    
    sheet_events.row_dimensions[row_idx].height = 50

# 调整列宽
sheet_events.column_dimensions['A'].width = 8
sheet_events.column_dimensions['B'].width = 35
sheet_events.column_dimensions['C'].width = 15
sheet_events.column_dimensions['D'].width = 15
sheet_events.column_dimensions['E'].width = 12
sheet_events.column_dimensions['F'].width = 12
sheet_events.column_dimensions['G'].width = 50
sheet_events.column_dimensions['H'].width = 35

# ===== Sheet 3: 粉丝情绪趋势 =====
sheet_emotion = wb.create_sheet('粉丝情绪趋势')
sheet_emotion['A1'] = '粉丝情绪趋势分析（2026年6月29日-7月5日）'
sheet_emotion.merge_cells('A1:F1')
set_cell_style(sheet_emotion['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_emotion.row_dimensions[1].height = 30

# 表头
headers = ['日期', '正面情绪占比', '负面情绪占比', '中性情绪占比', '品牌安全评分', '主要驱动事件']
for col, header in enumerate(headers, 1):
    cell = sheet_emotion.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_emotion.row_dimensions[2].height = 30

# 情绪数据（基于每日报告）
emotion_data = [
    ['2026-06-29', 50, 15, 35, 80, '新势力之夜倒计时，网传节目单混乱'],
    ['2026-06-30', 65, 10, 25, 82, '新势力之夜演出成功，表现获好评↑'],
    ['2026-07-01', 65, 10, 25, 82, '演出后正面反馈持续，黑西装造型获赞'],
    ['2026-07-02', 75, 10, 15, 78, '演出成功提振信心，但CP话题稀释个人品牌'],
    ['2026-07-03', 72, 10, 18, 76, '敷尔佳合作官宣，CP"吵架"话题突起'],
    ['2026-07-04', 65, 10, 25, 76, '敷尔佳TVC发布，CP风险持续关注'],
    ['2026-07-05', 75, 5, 20, 76, '敷尔佳壁纸发布，整体情绪稳定'],
]

for row_idx, row_data in enumerate(emotion_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_emotion.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置对齐和边框
        set_cell_style(cell, normal_font, None, Alignment(horizontal='center', vertical='center', wrap_text=True), thin_border)
        
        # 根据情绪占比设置条件格式
        if col_idx == 2 and isinstance(value, (int, float)) and value >= 70:
            cell.fill = success_fill
        elif col_idx == 3 and isinstance(value, (int, float)) and value >= 10:
            cell.fill = warning_fill
        elif col_idx == 5 and isinstance(value, (int, float)):
            if value >= 80:
                cell.fill = success_fill
            elif value >= 70:
                cell.fill = neutral_fill
            else:
                cell.fill = warning_fill
    
    sheet_emotion.row_dimensions[row_idx].height = 30

# 添加情绪趋势图
chart = BarChart()
chart.title = '粉丝情绪趋势（6月29日-7月5日）'
chart.style = 10
chart.y_axis.title = '情绪占比（%）'
chart.x_axis.title = '日期'
chart.height = 10
chart.width = 20

# 设置数据范围
data = Reference(sheet_emotion, min_col=2, min_row=2, max_row=9, max_col=4)
cats = Reference(sheet_emotion, min_col=1, min_row=3, max_row=9)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet_emotion.add_chart(chart, 'H3')

# 调整列宽
sheet_emotion.column_dimensions['A'].width = 15
sheet_emotion.column_dimensions['B'].width = 15
sheet_emotion.column_dimensions['C'].width = 15
sheet_emotion.column_dimensions['D'].width = 15
sheet_emotion.column_dimensions['E'].width = 15
sheet_emotion.column_dimensions['F'].width = 35

# ===== Sheet 4: 潜在风险汇总 =====
sheet_risks = wb.create_sheet('潜在风险汇总')
sheet_risks['A1'] = '潜在风险汇总与评估（2026年6月30日-7月5日）'
sheet_risks.merge_cells('A1:H1')
set_cell_style(sheet_risks['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_risks.row_dimensions[1].height = 30

# 表头
headers = ['风险等级', '风险类型', '风险描述', '影响范围', '发生概率', '应对建议', '责任方', '当前状态']
for col, header in enumerate(headers, 1):
    cell = sheet_risks.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_risks.row_dimensions[2].height = 30

# 风险数据（本周实际风险）
risks_data = [
    ['🔴 高风险', 'CP话题干扰品牌', '"左邓吵架"话题在敷尔佳宣发期出现，可能分散品牌传播效果', '敷尔佳合作效果、品牌形象', '高', '强化邓佳鑫个人标签，去CP化；引导评论区聚焦个人', '经纪团队+品牌方', '⚠️ 持续监控'],
    ['🔴 高风险', 'CP"恋情实锤"风险', '中国娱乐环境下，恋情传言若被坐实可能导致被雪藏', '艺人职业生涯', '中', '团队需明确态度，必要时进行舆论引导', '经纪团队', '⚠️ 持续监控'],
    ['🟡 中风险', 'TF_ING内部矛盾', '穆祉丞"躺照"争议持续发酵，粉丝阵营撕裂，可能波及邓佳鑫', '个人形象、粉丝关系', '中', '团队及时引导舆论，强调"团爱"；释放团体合作素材', '经纪团队', '⚠️ 需关注'],
    ['🟡 中风险', '竞品同期声量压制', '穆祉丞常州站（7/11-12）与敷尔佳直播期重叠，声量可能被分流', '敷尔佳合作效果', '高', '错峰运营；7/10前完成主要转化；增加邓佳鑫社媒更新', '经纪团队', '⚠️ 需关注'],
    ['🟡 中风险', '生日季过度营销', '7/23生日是粉丝经济黄金窗口，但过度营销可能引起路人反感', '品牌形象、路人好感', '中', '保持克制，可做公益/正能量活动；引导理性应援', '经纪团队+后援会', '⏳ 即将到来'],
    ['🟡 中风险', '运动会CP话题高发', 'TF家族夏日运动会（7/16-17）为无剧本互动，是CP话题高发场景', '个人品牌形象', '高', '提前准备舆情引导方案；突出个人高光时刻', '经纪团队', '⏳ 即将到来'],
    ['🟢 低风险', '历史争议重提', '解约风波（2025年12月）未被重提，但仍在搜索引擎有记录', '品牌形象', '低', '持续监控关键词，防止旧闻重炒', '公关团队', '✅ 可控'],
    ['🟢 低风险', '"每月下旬必爆"规律', '7月23日（生日）恰逢下旬，CP话题可能再度爆发', '个人品牌、粉丝关系', '中', '提前准备舆情应对方案；强化个人活动', '经纪团队', '⏳ 即将到来'],
]

for row_idx, row_data in enumerate(risks_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_risks.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据风险等级设置颜色
        if col_idx == 1:  # 风险等级列
            if '高' in str(value) and '🔴' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '低' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='008000'), success_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_risks.row_dimensions[row_idx].height = 60

# 调整列宽
sheet_risks.column_dimensions['A'].width = 15
sheet_risks.column_dimensions['B'].width = 20
sheet_risks.column_dimensions['C'].width = 40
sheet_risks.column_dimensions['D'].width = 25
sheet_risks.column_dimensions['E'].width = 15
sheet_risks.column_dimensions['F'].width = 40
sheet_risks.column_dimensions['G'].width = 15
sheet_risks.column_dimensions['H'].width = 15

# ===== Sheet 5: 营销建议和提醒 =====
sheet_marketing = wb.create_sheet('营销建议和提醒')
sheet_marketing['A1'] = '营销建议与行动计划（2026年7月5日更新）'
sheet_marketing.merge_cells('A1:H1')
set_cell_style(sheet_marketing['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_marketing.row_dimensions[1].height = 30

# 表头
headers = ['时间周期', '建议类型', '具体建议', '执行优先级', '预期效果', '负责方', '预算估算', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_marketing.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_marketing.row_dimensions[2].height = 30

# 营销建议数据（基于本周监测）
marketing_data = [
    ['紧急（今日-7/10）', '敷尔佳宣发全力配合', '今晚/明日持续发布TVC花絮，邓佳鑫本人社媒转发；发起#邓佳鑫敷尔佳品牌活力大使#话题', '🔴 最高', '提升品牌曝光，转化粉丝经济', '经纪团队+品牌方', '低', '7/3已启动，持续至7/10'],
    ['紧急（今日-7/10）', 'CP话题舆论引导', '"左邓吵架"话题若持续，及时发布邓佳鑫个人活动动态转移焦点；强化"唱作歌手"标签', '🔴 高', '避免CP干扰品牌合作', '经纪团队', '无', '敷尔佳合作期间特别重要'],
    ['短期（7/11-7/15）', '错峰竞品运营', '穆祉丞常州站期间（7/11-12），以静态内容为主；发布邓佳鑫个人舞台片段维持声量', '🟡 中', '避免声量被完全压制', '社媒运营团队', '低', '错峰发布'],
    ['短期（7/16-7/17）', '夏日运动会形象管理', '提前准备邓佳鑫个人高光时刻剪辑；防范CP话题过度发酵；突出"活力大使"定位', '🔴 高', '展示运动、活力形象', '经纪团队', '中', '澳门银河综艺馆'],
    ['中期（7/18-7/23）', '生日季营销策划', '21岁生日（7/23）可与敷尔佳合作推出"生日限定礼盒"；策划"成人礼"主题内容', '🔴 最高', '粉丝经济黄金窗口', '经纪团队+品牌方', '中', '需提前2周准备'],
    ['中期（7月内）', '《国宝寻音》播出配合', '人民日报出品的主流背书，提前准备传播素材；强调"文化担当""正能量偶像"', '🔴 高', '提升主流媒体好感度', '经纪团队+PR团队', '中', '播出时间待官方确认'],
    ['长期（8月起）', '个人品牌去CP化', '增加个人活动比重；强化"唱作歌手"标签；减少CP向内容', '🔴 高', '建立独特个人品牌', '经纪团队', '高', '需持续执行'],
    ['长期（8月起）', '商业价值提升', '维护敷尔佳合作，争取更多品牌合作；利用《国宝寻音》主流背书吸引同类品牌', '🟡 中', '提升商业价值', '经纪团队', '高', '需数据支撑'],
]

for row_idx, row_data in enumerate(marketing_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_marketing.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据优先级设置颜色
        if col_idx == 4:  # 执行优先级列
            if '最高' in str(value) or '🔴' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            else:
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center'))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_marketing.row_dimensions[row_idx].height = 60

# 调整列宽
sheet_marketing.column_dimensions['A'].width = 18
sheet_marketing.column_dimensions['B'].width = 20
sheet_marketing.column_dimensions['C'].width = 50
sheet_marketing.column_dimensions['D'].width = 15
sheet_marketing.column_dimensions['E'].width = 25
sheet_marketing.column_dimensions['F'].width = 20
sheet_marketing.column_dimensions['G'].width = 15
sheet_marketing.column_dimensions['H'].width = 20

# ===== Sheet 6: 下周重点关注事项 =====
sheet_nextweek = wb.create_sheet('下周重点关注事项')
sheet_nextweek['A1'] = '下周重点关注事项（2026年7月6日-7月12日）'
sheet_nextweek.merge_cells('A1:H1')
set_cell_style(sheet_nextweek['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_nextweek.row_dimensions[1].height = 30

# 表头
headers = ['日期', '事件', '重要性', '关注要点', '应对措施', '负责方', '预期结果', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_nextweek.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_nextweek.row_dimensions[2].height = 30

# 下周关注事项数据
nextweek_data = [
    ['2026-07-06', '敷尔佳TVC花絮预热2发布', '⭐⭐⭐⭐', '社媒反响、评论区情绪、CP话题是否干扰', '及时监测，引导舆论聚焦邓佳鑫个人', '社媒运营团队', '维持宣发热度', '19:23发布'],
    ['2026-07-08', '敷尔佳直播活动官宣', '⭐⭐⭐⭐⭐', '直播时间、平台、互动机制公布', '提前准备粉丝互动机制（抽奖、限定福利）', '品牌方+经纪团队', '为直播预热', '10:23官宣'],
    ['2026-07-09', '敷尔佳直播购买攻略发布', '⭐⭐⭐⭐', '粉丝购买指引、优惠信息', '确保信息准确，引导理性消费', '品牌方', '提升转化率', '19:23发布'],
    ['2026-07-10', '敷尔佳直播倒计时海报', '⭐⭐⭐⭐', '最后冲刺，粉丝期待值管理', '强化紧迫感，促进预售', '品牌方', '直播预热峰值', '15:23发布'],
    ['2026-07-11', '穆祉丞常州站演唱会（Day1）', '⭐⭐⭐', '竞品声量、粉丝注意力分配', '以静态内容为主，避免重大宣发', '社媒运营团队', '避免声量被完全覆盖', '⚠️ 竞品同期'],
    ['2026-07-12', '穆祉丞常州站演唱会（Day2）', '⭐⭐⭐', '同上；门票今日开售', '观察购票热度，评估对邓佳鑫声量影响', '社媒运营团队', '调整下周策略', '⚠️ 竞品同期'],
    ['2026-07-13', '下周策略调整', '⭐⭐⭐⭐', '敷尔佳合作效果评估、CP话题控制效果', '总结经验，调整夏日运动会策略', '经纪团队', '优化后续运营', '周一策略会'],
]

for row_idx, row_data in enumerate(nextweek_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_nextweek.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据重要性设置颜色
        if col_idx == 3:  # 重要性列
            if '⭐⭐⭐⭐⭐' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center'))
            elif '⭐⭐⭐⭐' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center'))
            else:
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center'))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_nextweek.row_dimensions[row_idx].height = 50

# 调整列宽
sheet_nextweek.column_dimensions['A'].width = 15
sheet_nextweek.column_dimensions['B'].width = 30
sheet_nextweek.column_dimensions['C'].width = 15
sheet_nextweek.column_dimensions['D'].width = 35
sheet_nextweek.column_dimensions['E'].width = 30
sheet_nextweek.column_dimensions['F'].width = 20
sheet_nextweek.column_dimensions['G'].width = 25
sheet_nextweek.column_dimensions['H'].width = 20

# ===== Sheet 7: 敷尔佳合作专题 =====
sheet_fuhaner = wb.create_sheet('敷尔佳合作专题')
sheet_fuhaner['A1'] = '敷尔佳×邓佳鑫品牌合作专题（2026年7月）'
sheet_fuhaner.merge_cells('A1:H1')
set_cell_style(sheet_fuhaner['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_fuhaner.row_dimensions[1].height = 30

# 合作概述
sheet_fuhaner['A3'] = '合作概述'
set_cell_style(sheet_fuhaner['A3'], subtitle_font, None, Alignment(horizontal='left', vertical='center'))
sheet_fuhaner.row_dimensions[3].height = 25

overview_data = [
    ('合作身份', '敷尔佳品牌活力大使'),
    ('合作时间', '2026年5月官宣，7月正式宣发'),
    ('合作定位', '打破"仅属于问题肌肤"刻板印象，向年轻群体传递"精简护肤"理念'),
    ('定制联名礼盒', '已打造，线上线下多元传播'),
    ('开售24h战报', '销售额突破240w+'),
    ('微博粉丝', '邓佳鑫554.6万，与目标受众匹配度高'),
]

for i, (key, value) in enumerate(overview_data):
    row = 4 + i
    sheet_fuhaner[f'A{row}'] = key
    sheet_fuhaner[f'B{row}'] = value
    set_cell_style(sheet_fuhaner[f'A{row}'], Font(name='Arial', bold=True, size=10), None, Alignment(horizontal='right', vertical='center'))
    set_cell_style(sheet_fuhaner[f'B{row}'], normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True))
    sheet_fuhaner.row_dimensions[row].height = 20

# 宣发时间表
sheet_fuhaner['A12'] = '完整宣发时间表（粉丝站曝光版）'
set_cell_style(sheet_fuhaner['A12'], subtitle_font, None, Alignment(horizontal='left', vertical='center'))
sheet_fuhaner.row_dimensions[12].height = 25

schedule_headers = ['日期', '时间', '内容', '发布渠道', '状态', '备注']
for col, header in enumerate(schedule_headers, 1):
    cell = sheet_fuhaner.cell(row=13, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center'))
    sheet_fuhaner.row_dimensions[13].height = 25

schedule_data = [
    ['2026-07-03', '19:23', 'TVC花絮预热1', '品牌微博/小红书/视频号/抖音', '✅ 已发布', '社媒反响正面'],
    ['2026-07-05', '15:23', '限定壁纸发布', '品牌微博/小红书', '⏳ 今日发布', '当前时间09:50，约5.5小时后'],
    ['2026-07-06', '19:23', 'TVC花絮预热2', '品牌微博/小红书/视频号/抖音', '⏳ 待发布', ''],
    ['2026-07-08', '10:23', '直播活动官宣', '品牌&艺人微博/小红书', '⏳ 待发布', '转化关键节点'],
    ['2026-07-09', '19:23', '直播购买攻略', '全平台', '⏳ 待发布', ''],
    ['2026-07-10', '15:23', '直播倒计时海报', '全平台', '⏳ 待发布', ''],
]

for row_idx, row_data in enumerate(schedule_data, 14):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_fuhaner.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据状态设置颜色
        if col_idx == 5:  # 状态列
            if '✅' in str(value):
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center'))
            elif '⏳' in str(value):
                set_cell_style(cell, normal_font, neutral_fill, Alignment(horizontal='center', vertical='center'))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='center', vertical='center', wrap_text=True))
        
        cell.border = thin_border
    
    sheet_fuhaner.row_dimensions[row_idx].height = 30

# 合作风险提示
sheet_fuhaner['A22'] = '合作期间风险提示'
set_cell_style(sheet_fuhaner['A22'], subtitle_font, None, Alignment(horizontal='left', vertical='center'))
sheet_fuhaner.row_dimensions[22].height = 25

risk_data = [
    ('CP话题干扰', '护肤品牌对艺人形象要求较高，CP传闻可能影响品牌调性', '🔴 高风险', '强化个人标签，去CP化'),
    ('竞品同期声量', '穆祉丞常州站（7/11-12）与直播期重叠', '🟡 中风险', '错峰运营，7/10前完成主要转化'),
    ('粉丝内耗', 'CP粉与唯粉矛盾在宣发期爆发', '🟡 中风险', '引导舆论聚焦邓佳鑫个人魅力'),
    ('销售不及预期', '粉丝购买力受竞品演唱会门票分流', '🟢 低风险', '提前准备限时优惠，促进转化'),
]

for i, (risk, desc, level, action) in enumerate(risk_data):
    row = 23 + i
    sheet_fuhaner[f'A{row}'] = risk
    sheet_fuhaner[f'B{row}'] = desc
    sheet_fuhaner[f'C{row}'] = level
    sheet_fuhaner[f'D{row}'] = action
    
    set_cell_style(sheet_fuhaner[f'A{row}'], Font(name='Arial', bold=True, size=10), None, Alignment(horizontal='left', vertical='center'))
    set_cell_style(sheet_fuhaner[f'B{row}'], normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True))
    
    if '高' in level and '🔴' in level:
        set_cell_style(sheet_fuhaner[f'C{row}'], Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center'))
    elif '中' in level:
        set_cell_style(sheet_fuhaner[f'C{row}'], Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center'))
    else:
        set_cell_style(sheet_fuhaner[f'C{row}'], Font(name='Arial', bold=True, size=10, color='008000'), success_fill, Alignment(horizontal='center', vertical='center'))
    
    set_cell_style(sheet_fuhaner[f'D{row}'], normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True))
    
    sheet_fuhaner.row_dimensions[row].height = 40

# 调整列宽
sheet_fuhaner.column_dimensions['A'].width = 20
sheet_fuhaner.column_dimensions['B'].width = 30
sheet_fuhaner.column_dimensions['C'].width = 15
sheet_fuhaner.column_dimensions['D'].width = 30
sheet_fuhaner.column_dimensions['E'].width = 20

# ===== Sheet 8: 竞品动态对比 =====
sheet_competitors = wb.create_sheet('竞品动态对比')
sheet_competitors['A1'] = '竞品动态对比分析（2026年6月30日-7月5日）'
sheet_competitors.merge_cells('A1:H1')
set_cell_style(sheet_competitors['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_competitors.row_dimensions[1].height = 30

# 表头
headers = ['竞品姓名', '最新动态', '热度等级', '对邓佳鑫的影响', '应对策略', '粉丝规模对比', '品牌安全对比', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_competitors.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_competitors.row_dimensions[2].height = 30

# 竞品数据
competitors_data = [
    ['穆祉丞', '「局部宇宙」巡演诸暨站已结束；常州站7/11-12；"躺照"争议引发内部矛盾讨论', '⭐⭐⭐⭐⭐', '演唱会场UGC持续发酵；常州站与敷尔佳直播期重叠，声量可能被分流', '错峰运营；7/10前完成主要转化；增加邓佳鑫社媒更新频次', '高会粉丝14万+，生日周边销售破八位数', '内部矛盾风险上升', '🔴 主要竞品'],
    ['左航', '"左邓"CP话题核心另一方；"左邓吵架"视频获4.9万赞', '⭐⭐⭐⭐', 'CP话题可能稀释邓佳鑫个人品牌；敷尔佳合作期间需特别防范', '评估公司态度；强化个人活动；必要时舆论引导', '粉丝规模较大', '受CP话题影响', '⚠️ CP关联'],
    ['张子墨', '个人专辑（6/30上线）获好评；新专辑先行曲《无以言表的日子里》治愈曲风', '⭐⭐⭐', '与邓佳鑫同为TF_ING主唱担当，良性竞争', '打造差异化内容；突出邓佳鑫独特优势', '粉丝规模中等', '平稳', '竞争相对温和'],
    ['黄朔', '黑西装红毯造型持续获赞；新势力之夜同台', '⭐⭐⭐', '组合形式登台，有助于提升TF_ING整体声量', '借助组合曝光，提升个人知名度', '粉丝规模较小', '平稳', '竞争相对温和'],
    ['童禹坤', 'TF_ING成员，参与东方风云榜', '⭐⭐', '粉丝对比各成员资源分配', '提升邓佳鑫个人活动质量', '粉丝规模中等', '平稳', '竞争相对温和'],
]

for row_idx, row_data in enumerate(competitors_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_competitors.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_competitors.row_dimensions[row_idx].height = 60

# 调整列宽
sheet_competitors.column_dimensions['A'].width = 15
sheet_competitors.column_dimensions['B'].width = 40
sheet_competitors.column_dimensions['C'].width = 15
sheet_competitors.column_dimensions['D'].width = 30
sheet_competitors.column_dimensions['E'].width = 30
sheet_competitors.column_dimensions['F'].width = 20
sheet_competitors.column_dimensions['G'].width = 20
sheet_competitors.column_dimensions['H'].width = 20

# 保存Excel文件
output_file = '邓佳鑫监测周报_20260630-20260705.xlsx'
wb.save(output_file)
print(f'✅ Excel周报已生成：{output_file}')
print(f'📊 包含8个工作表：')
print(f'  1. 封面')
print(f'  2. 本周热点事件（7个热点）')
print(f'  3. 粉丝情绪趋势（7天数据+图表）')
print(f'  4. 潜在风险汇总（8个风险项）')
print(f'  5. 营销建议和提醒（8条建议）')
print(f'  6. 下周重点关注事项（7个事项）')
print(f'  7. 敷尔佳合作专题（合作详情+时间表+风险）')
print(f'  8. 竞品动态对比（5个竞品）')
