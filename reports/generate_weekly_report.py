#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
邓佳鑫监测周报Excel生成脚本
生成时间：2026年6月28日
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# 创建Excel工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认sheet

# 定义样式
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14, color='1F3864')
subtitle_font = Font(name='Arial', bold=True, size=12, color='2E75B6')
normal_font = Font(name='Arial', size=10)
warning_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
neutral_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

# ===== Sheet 1: 封面 =====
sheet_cover = wb.create_sheet('封面')
sheet_cover.merge_cells('A1:H1')
sheet_cover['A1'] = '邓佳鑫全网监测周报'
set_cell_style(sheet_cover['A1'], title_font, None, Alignment(horizontal='center', vertical='center'))
sheet_cover['A1'].font = Font(name='Arial', bold=True, size=20, color='1F3864')
sheet_cover.row_dimensions[1].height = 40

sheet_cover.merge_cells('A3:H3')
sheet_cover['A3'] = '2026年6月22日 - 6月28日（第26周）'
set_cell_style(sheet_cover['A3'], subtitle_font, None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[3].height = 30

sheet_cover.merge_cells('A5:H5')
sheet_cover['A5'] = '报告生成时间：2026年6月28日'
set_cell_style(sheet_cover['A5'], normal_font, None, Alignment(horizontal='center', vertical='center'))

sheet_cover.merge_cells('A7:H7')
sheet_cover['A7'] = '生成工具：WorkBuddy 自动化监测系统'
set_cell_style(sheet_cover['A7'], normal_font, None, Alignment(horizontal='center', vertical='center'))

# 添加品牌安全评分
sheet_cover.merge_cells('A10:C10')
sheet_cover['A10'] = '品牌安全评分'
set_cell_style(sheet_cover['A10'], subtitle_font, None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[10].height = 25

sheet_cover.merge_cells('A11:C11')
sheet_cover['A11'] = '82/100'
set_cell_style(sheet_cover['A11'], Font(name='Arial', bold=True, size=24, color='FF0000'), None, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[11].height = 40

sheet_cover.merge_cells('A13:H13')
sheet_cover['A13'] = '⚠️ 本周评级：中等风险（CP话题热搜，粉丝情绪分化）'
set_cell_style(sheet_cover['A13'], Font(name='Arial', bold=True, size=12, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center'))
sheet_cover.row_dimensions[13].height = 30

# 调整封面sheet列宽
sheet_cover.column_dimensions['A'].width = 20
sheet_cover.column_dimensions['B'].width = 20
sheet_cover.column_dimensions['C'].width = 20

# ===== Sheet 2: 本周热点事件 =====
sheet_events = wb.create_sheet('本周热点事件')
sheet_events['A1'] = '本周热点事件列表（2026年6月22日-6月28日）'
sheet_events.merge_cells('A1:H1')
set_cell_style(sheet_events['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_events.row_dimensions[1].height = 30

# 表头
headers = ['序号', '事件名称', '热度等级', '发生时间', '情绪倾向', '风险等级', '相关链接', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_events.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_events.row_dimensions[2].height = 30

# 热点事件数据
events_data = [
    [1, '《国宝寻音》综艺录制完成', '⭐⭐⭐⭐⭐', '2026-06-21', '正面', '低风险', 'https://www.shobserver.com/staticsg/res/html/web/newsDetail.html?id=1114828&sid=11', '人民日报出品，7月播出'],
    [2, '敷尔佳联名礼盒持续发酵', '⭐⭐⭐⭐', '2026-06-11-12', '正面', '低风险', 'https://hb.dzwww.com/p/p6ucAZYe9Gc.html', '品牌年轻化成功案例'],
    [3, '"左邓"CP热搜', '⭐⭐⭐⭐⭐', '2026-06-23', '两极分化', '高风险', 'https://ent.sina.cn/2026-06-24/detail-iniemshn6544331.d.html', '粉丝圈两极分化'],
    [4, '东方风云榜新势力之夜倒计时', '⭐⭐⭐⭐', '2026-06-30', '正面', '中风险', 'https://baike.baidu.com/item/看东方·东方风云榜新势力之夜/67719995', '邓佳鑫25分钟微个唱'],
    [5, '穆祉丞"局部宇宙"演唱会', '⭐⭐⭐⭐', '2026-06-27-28', '中性', '中风险', 'https://ent.sina.cn/2026-06-27/detail-inieveqv6936639.d.html', '竞品动态，声量分流'],
    [6, '邓佳鑫武汉汽水音乐节表现', '⭐⭐⭐', '2026-06-23', '正面', '低风险', 'https://www.iqiyi.com/v_13kk4djtv4k.html', '全开麦表现稳定'],
    [7, 'TF家族运动会倒计时20天', '⭐⭐⭐', '2026-07-16-17', '期待', '中风险', 'https://baijiahao.baidu.com/s?id=1869029202513808718', '官方未官宣，粉丝催更']
]

for row_idx, row_data in enumerate(events_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_events.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据风险等级设置颜色
        if col_idx == 6:  # 风险等级列
            if '高' in str(value):
                set_cell_style(cell, normal_font, warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, normal_font, neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '低' in str(value):
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True))
        
        cell.border = thin_border
    
    sheet_events.row_dimensions[row_idx].height = 40

# 调整列宽
sheet_events.column_dimensions['A'].width = 8
sheet_events.column_dimensions['B'].width = 30
sheet_events.column_dimensions['C'].width = 15
sheet_events.column_dimensions['D'].width = 15
sheet_events.column_dimensions['E'].width = 12
sheet_events.column_dimensions['F'].width = 12
sheet_events.column_dimensions['G'].width = 50
sheet_events.column_dimensions['H'].width = 30

# ===== Sheet 3: 粉丝情绪趋势 =====
sheet_emotion = wb.create_sheet('粉丝情绪趋势')
sheet_emotion['A1'] = '粉丝情绪趋势分析（2026年6月22日-6月28日）'
sheet_emotion.merge_cells('A1:F1')
set_cell_style(sheet_emotion['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_emotion.row_dimensions[1].height = 30

# 表头
headers = ['日期', '正面情绪占比', '负面情绪占比', '中性情绪占比', '主要驱动事件', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_emotion.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_emotion.row_dimensions[2].height = 30

# 情绪数据
emotion_data = [
    ['2026-06-22', 80, 5, 15, '《国宝寻音》录制完成，敷尔佳合作持续', '品牌安全评分88/100'],
    ['2026-06-23', 80, 5, 15, '平稳期，无重大新动态', '品牌安全评分88/100'],
    ['2026-06-24', 65, 15, 20, '"左邓"CP热搜，粉丝两极分化', '品牌安全评分82/100↓'],
    ['2026-06-25', 65, 15, 20, 'CP话题持续发酵，微博"有猫饼"更新', '情绪分化明显'],
    ['2026-06-26', 55, 25, 20, 'CP热搜余波，唯粉与CP粉矛盾', '负面情绪上升'],
    ['2026-06-27', 45, 25, 30, '穆祉丞演唱会开幕，声量分流', '正面情绪下降至45%'],
    ['2026-06-28', 55, 15, 30, '穆祉丞公开感谢邓佳鑫，CP话题升温', '情绪略有回升']
]

for row_idx, row_data in enumerate(emotion_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_emotion.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置对齐和边框
        set_cell_style(cell, normal_font, None, Alignment(horizontal='center', vertical='center', wrap_text=True), thin_border)
        
        # 根据情绪占比设置条件格式（简化版，通过颜色标识）
        if col_idx in [2, 3, 4]:  # 情绪占比列
            if col_idx == 2 and isinstance(value, (int, float)) and value >= 60:
                cell.fill = success_fill
            elif col_idx == 3 and isinstance(value, (int, float)) and value >= 20:
                cell.fill = warning_fill
    
    sheet_emotion.row_dimensions[row_idx].height = 30

# 添加情绪趋势图（使用条形图）
chart = BarChart()
chart.title = '粉丝情绪趋势（6月22日-6月28日）'
chart.style = 10
chart.y_axis.title = '情绪占比（%）'
chart.x_axis.title = '日期'
chart.height = 10
chart.width = 20

# 设置数据范围
data = Reference(sheet_emotion, min_col=2, min_row=2, max_row=9, max_col=4)
cats = Reference(sheet_emotion, min_col=1, min_row=3, max_row=9)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

sheet_emotion.add_chart(chart, 'H3')

# 调整列宽
sheet_emotion.column_dimensions['A'].width = 15
sheet_emotion.column_dimensions['B'].width = 15
sheet_emotion.column_dimensions['C'].width = 15
sheet_emotion.column_dimensions['D'].width = 15
sheet_emotion.column_dimensions['E'].width = 30
sheet_emotion.column_dimensions['F'].width = 20

# ===== Sheet 4: 潜在风险汇总 =====
sheet_risks = wb.create_sheet('潜在风险汇总')
sheet_risks['A1'] = '潜在风险汇总与评估'
sheet_risks.merge_cells('A1:H1')
set_cell_style(sheet_risks['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_risks.row_dimensions[1].height = 30

# 表头
headers = ['风险等级', '风险类型', '风险描述', '影响范围', '发生概率', '应对建议', '责任方', '状态']
for col, header in enumerate(headers, 1):
    cell = sheet_risks.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_risks.row_dimensions[2].height = 30

# 风险数据
risks_data = [
    ['🔴 高风险', 'CP捆绑风险', '"左邓"CP热搜频率过高（每月下旬一次），公司已注意到但未表态', '品牌安全、粉丝凝聚力、个人形象', '高', '密切关注公司态度，准备危机公关预案，强化个人成绩宣传', '经纪团队', '⚠️ 持续监控'],
    ['🔴 高风险', '热搜规律性争议', '"每月下旬必爆一次"规律被公开讨论，可能引发舆论反弹', '公众形象、路人好感', '中', '适当降低热搜频率，避免规律化，引导粉丝低调', '经纪团队', '⚠️ 持续监控'],
    ['🟡 中风险', '粉丝情绪分化', '唯粉与CP粉矛盾可能激化，影响粉丝购买力', '商务转化、粉丝活动', '中', '通过优质内容凝聚粉丝，避免过度刺激任一群体', '后援会', '⚠️ 需关注'],
    ['🟡 中风险', '竞品流量分流', '穆祉丞音综首秀、演唱会可能吸引部分关注', '曝光度、话题量', '高', '《国宝寻音》7月播出作为反击节点，提前准备播出后的社媒互动', '经纪团队', '⚠️ 需关注'],
    ['🟡 中风险', '6月30日表现压力', '25分钟微个唱是重要个人展示机会，表现直接影响后续资源', '专业形象、后续资源', '中', '充分排练，确保现场表现稳定，准备应急预案', '邓佳鑫本人', '⏳ 即将发生'],
    ['🟢 低风险', '历史争议重提', '与时代峰峻的合约纠纷仍在搜索引擎有记录', '品牌形象', '低', '持续监控"时代峰峻"+"邓佳鑫"关键词，防止旧闻重炒', '公关团队', '✅ 可控'],
    ['🟢 低风险', '品牌形象关联风险', 'CP话题过度蔓延，可能影响邓佳鑫"独立音乐人"形象', '商业价值', '中', '增加正面、专业向内容，提升个人品牌形象', '经纪团队', '✅ 可控']
]

for row_idx, row_data in enumerate(risks_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_risks.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据风险等级设置颜色
        if col_idx == 1:  # 风险等级列
            if '高' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '低' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='008000'), success_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_risks.row_dimensions[row_idx].height = 50

# 调整列宽
sheet_risks.column_dimensions['A'].width = 15
sheet_risks.column_dimensions['B'].width = 20
sheet_risks.column_dimensions['C'].width = 40
sheet_risks.column_dimensions['D'].width = 25
sheet_risks.column_dimensions['E'].width = 15
sheet_risks.column_dimensions['F'].width = 40
sheet_risks.column_dimensions['G'].width = 15
sheet_risks.column_dimensions['H'].width = 15

# ===== Sheet 5: 营销建议 =====
sheet_marketing = wb.create_sheet('营销建议')
sheet_marketing['A1'] = '营销建议与行动计划'
sheet_marketing.merge_cells('A1:H1')
set_cell_style(sheet_marketing['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_marketing.row_dimensions[1].height = 30

# 表头
headers = ['时间周期', '建议类型', '具体建议', '执行优先级', '预期效果', '负责方', '预算估算', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_marketing.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_marketing.row_dimensions[2].height = 30

# 营销建议数据
marketing_data = [
    ['短期（1周内）', '新势力之夜预热', '提前准备品牌联动内容（演出妆容使用敷尔佳产品，发布后台花絮）', '🔴 高', '提升演出热度，增加品牌曝光', '经纪团队+品牌方', '低', '6月30日执行'],
    ['短期（1周内）', 'CP话题引导', '若公司不鼓励CP捆绑，建议低调处理，不直接回应；若接受，可适度引导至"音乐合作伙伴"方向', '🟡 中', '避免粉丝内战，维护品牌形象', '经纪团队', '无', '需评估公司态度'],
    ['短期（1周内）', '应对竞品流量', '6月27-28日减少重大发布；6月29日集中释放邓佳鑫正面内容', '🟡 中', '避免被竞品流量覆盖', '社媒运营团队', '低', '实时监控穆祉丞演唱会UGC'],
    ['中期（1个月内）', '《国宝寻音》播出配合', '提前准备邓佳鑫在节目中的高光片段，用于社媒传播；结合"传统文化+现代音乐"标签', '🔴 高', '提升主流媒体好感度，吸引路人粉', '经纪团队+PR团队', '中', '7月播出，需提前2周准备'],
    ['中期（1个月内）', '加强抖音/小红书投放', '《国宝寻音》录制花絮、国宝乐器科普、城市文化内容；敷尔佳礼盒开箱、使用心得等UGC内容引导', '🟡 中', '扩大年轻受众，提升商业价值', '社媒运营团队', '中', '需制作高质量短视频内容'],
    ['中期（1个月内）', '粉丝管理', '关注唯粉与CP粉矛盾，必要时进行适度引导；防止粉丝内部冲突影响品牌形象', '🟡 中', '维护粉丝圈和谐，防止负面舆论外溢', '后援会+经纪团队', '低', '需后援会配合'],
    ['长期（3个月内）', '巩固"文化和音乐双重属性"定位', '《国宝寻音》建立了文化综艺赛道优势，可延续此方向，参与更多文化音乐类节目', '🔴 高', '建立独特标识，减少CP话题干扰', '经纪团队', '高', '需筛选合适节目'],
    ['长期（3个月内）', '建立更系统的粉丝管理机制', '避免粉丝圈层冲突外溢到品牌层面', '🟡 中', '提升粉丝购买力，维护品牌形象', '经纪团队+后援会', '中', '需制定粉丝管理手册'],
    ['长期（3个月内）', '考虑参与更多音乐类综艺', '巩固"三代第一vocal"定位', '🔴 高', '提升专业认可度，吸引路人和业内人士', '经纪团队', '高', '需筛选合适音综']
]

for row_idx, row_data in enumerate(marketing_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_marketing.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据优先级设置颜色
        if col_idx == 4:  # 执行优先级列
            if '高' in str(value) and '🔴' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            elif '中' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
            else:
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center'))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_marketing.row_dimensions[row_idx].height = 60

# 调整列宽
sheet_marketing.column_dimensions['A'].width = 18
sheet_marketing.column_dimensions['B'].width = 20
sheet_marketing.column_dimensions['C'].width = 50
sheet_marketing.column_dimensions['D'].width = 15
sheet_marketing.column_dimensions['E'].width = 25
sheet_marketing.column_dimensions['F'].width = 20
sheet_marketing.column_dimensions['G'].width = 15
sheet_marketing.column_dimensions['H'].width = 20

# ===== Sheet 6: 下周重点关注事项 =====
sheet_nextweek = wb.create_sheet('下周重点关注事项')
sheet_nextweek['A1'] = '下周重点关注事项（2026年6月29日-7月5日）'
sheet_nextweek.merge_cells('A1:H1')
set_cell_style(sheet_nextweek['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_nextweek.row_dimensions[1].height = 30

# 表头
headers = ['日期', '事件', '重要性', '关注要点', '应对措施', '负责方', '预期结果', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_nextweek.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_nextweek.row_dimensions[2].height = 30

# 下周关注事项数据
nextweek_data = [
    ['2026-06-29', '东方风云榜新势力之夜预热', '⭐⭐⭐⭐', '社媒预热内容发布，粉丝互动', '提前准备现场素材，演出后快速扩散社媒', '社媒运营团队', '提升演出热度，增加曝光', '倒计时1天'],
    ['2026-06-30', '东方风云榜新势力之夜演出', '⭐⭐⭐⭐⭐', '25分钟微个唱表现，粉丝反响', '确保现场发挥稳定，快速推送精彩片段', '邓佳鑫本人+经纪团队', '提升专业形象，增加粉丝忠诚度', '🔴 关键节点'],
    ['2026-07-01', '新势力之夜后续传播', '⭐⭐⭐⭐', '演出cut传播，粉丝UGC内容', '官方发布高质量舞台视频，引导粉丝创作UGC', '社媒运营团队', '延续演出热度，吸引路人关注', '演出后1-2天是关键期'],
    ['2026-07-02', '"左邓"CP话题监控', '⭐⭐⭐', '是否继续高热度，公司是否介入', '每日监测CP话题热度，及时调整策略', '公关团队', '避免CP话题过度蔓延', '需密切关注公司态度'],
    ['2026-07-03', '竞品动态监控', '⭐⭐⭐', '穆祉丞演唱会后续UGC，其他TF_ING成员动态', '分析竞品UGC内容，找出差异化优势', '社媒运营团队', '制定差异化内容策略', '持续监控'],
    ['2026-07-04', '敷尔佳合作后续', '⭐⭐⭐⭐', '联名礼盒销售数据，粉丝反馈', '分析销售数据，规划后续合作内容', '品牌方+经纪团队', '提升商业价值，吸引同类品牌', '需与品牌方保持沟通'],
    ['2026-07-05', '《国宝寻音》播出倒计时', '⭐⭐⭐⭐', '播出时间官宣，预热内容准备', '提前准备传播素材，强调"文化传承"人设', '经纪团队+PR团队', '为主流媒体曝光做准备', '预计7月中旬播出']
]

for row_idx, row_data in enumerate(nextweek_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_nextweek.cell(row=row_idx, column=col_idx, value=value)
        
        # 根据重要性设置颜色
        if col_idx == 3:  # 重要性列
            if '⭐⭐⭐⭐⭐' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF0000'), warning_fill, Alignment(horizontal='center', vertical='center'))
            elif '⭐⭐⭐⭐' in str(value):
                set_cell_style(cell, Font(name='Arial', bold=True, size=10, color='FF8C00'), neutral_fill, Alignment(horizontal='center', vertical='center'))
            else:
                set_cell_style(cell, normal_font, success_fill, Alignment(horizontal='center', vertical='center'))
        else:
            set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_nextweek.row_dimensions[row_idx].height = 50

# 调整列宽
sheet_nextweek.column_dimensions['A'].width = 15
sheet_nextweek.column_dimensions['B'].width = 25
sheet_nextweek.column_dimensions['C'].width = 15
sheet_nextweek.column_dimensions['D'].width = 30
sheet_nextweek.column_dimensions['E'].width = 30
sheet_nextweek.column_dimensions['F'].width = 20
sheet_nextweek.column_dimensions['G'].width = 25
sheet_nextweek.column_dimensions['H'].width = 20

# ===== Sheet 7: 竞品动态对比 =====
sheet_competitors = wb.create_sheet('竞品动态对比')
sheet_competitors['A1'] = '竞品动态对比分析（2026年6月22日-6月28日）'
sheet_competitors.merge_cells('A1:H1')
set_cell_style(sheet_competitors['A1'], title_font, None, Alignment(horizontal='left', vertical='center'))
sheet_competitors.row_dimensions[1].height = 30

# 表头
headers = ['竞品姓名', '最新动态', '热度等级', '对邓佳鑫的影响', '应对策略', '粉丝规模对比', '品牌安全对比', '备注']
for col, header in enumerate(headers, 1):
    cell = sheet_competitors.cell(row=2, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, Alignment(horizontal='center', vertical='center', wrap_text=True))
    sheet_competitors.row_dimensions[2].height = 30

# 竞品数据
competitors_data = [
    ['穆祉丞', '《天赐的声音》第七季音综首秀；"局部宇宙"演唱会诸暨站（6月27-28日）；父亲节888元红包热搜', '⭐⭐⭐⭐⭐', '演唱会场分流TF家族粉丝注意力，压制邓佳鑫声量', '6月29日-7月1日集中发布邓佳鑫正面内容；借势《国宝寻音》差异化定位', '高会粉丝约10万+，稳居三代首位', '相对平稳，暂无重大负面', '主要竞品，需重点监控'],
    ['左航', '"左邓"CP话题核心另一方，近期无独立负面', '⭐⭐⭐⭐', 'CP话题可能引发公司介入，影响个人商务', '评估公司态度，明确CP营销策略；强化个人成绩宣传', '粉丝规模中等', '受CP话题影响，个人品牌被稀释', 'CP关联，需关注公司态度'],
    ['童禹坤', '专场嘉宾为余宇涵，粉丝讨论各成员嘉宾配置', '⭐⭐⭐', '粉丝对比各成员嘉宾配置，可能影响邓佳鑫粉丝满意度', '提升邓佳鑫个人活动质量，增加粉丝互动', '粉丝规模中等', '平稳', '竞争相对温和'],
    ['黄朔、张子墨', '东方风云榜新势力之夜同台，无重大动态', '⭐⭐', '组合形式登台，有助于提升TF_ING整体声量', '借助组合曝光，提升个人知名度', '粉丝规模较小', '平稳', '竞争相对温和']
]

for row_idx, row_data in enumerate(competitors_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = sheet_competitors.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, normal_font, None, Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    
    sheet_competitors.row_dimensions[row_idx].height = 60

# 调整列宽
sheet_competitors.column_dimensions['A'].width = 15
sheet_competitors.column_dimensions['B'].width = 40
sheet_competitors.column_dimensions['C'].width = 15
sheet_competitors.column_dimensions['D'].width = 30
sheet_competitors.column_dimensions['E'].width = 30
sheet_competitors.column_dimensions['F'].width = 20
sheet_competitors.column_dimensions['G'].width = 20
sheet_competitors.column_dimensions['H'].width = 20

# 保存Excel文件
output_file = '邓佳鑫监测周报_20260622-20260628.xlsx'
wb.save(output_file)
print(f'✅ Excel周报已生成：{output_file}')
print(f'📊 包含7个工作表：')
print(f'  1. 封面')
print(f'  2. 本周热点事件（7个热点）')
print(f'  3. 粉丝情绪趋势（7天数据+图表）')
print(f'  4. 潜在风险汇总（7个风险项）')
print(f'  5. 营销建议（9条建议）')
print(f'  6. 下周重点关注事项（7个事项）')
print(f'  7. 竞品动态对比（4个竞品）')
